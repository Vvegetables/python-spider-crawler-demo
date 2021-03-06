#coding=utf-8
from openpyxl.workbook.workbook import Workbook
from openpyxl.reader.excel import load_workbook

def excel_read(path):
    wb = load_workbook(path)
    sheet = wb.active
    return sheet
#     max_row = sheet.max_row
#     
#     for i in range(1,max_row+1):
#         if not sheet.cell(row=i,column=3).value:
#             continue
#         flag = UrlRedis.add_url(sheet.cell(row=i,column=3).value)
#         if flag:
#             continue
        
def get_excel_data(path=None,sheet=None,row_id=1):
    sheet = sheet or(excel_read(path) if path else None)
    container = []
    max_column = sheet.max_column      
    for i in range(row_id,max_column+1):
        if not sheet.cell(row=1,column=i).value:
            continue
        container.append(sheet.cell(row=1,column=i).value)
    return container
    
# def get_single_row_data(path=None,sheet=None,row_id=1):
#     get_title()

def get_single_column_data(sheet,colum_id):
    container = []
    max_row = sheet.max_row      
    for i in range(3,max_row+1):
        if not sheet.cell(row=i,column=colum_id).value:
            continue
        container.append(sheet.cell(row=i,column=colum_id).value)
    return container

def get_excel_cell_data(sheet,row,column):
    v = sheet.cell(row=row,column=column).value
    return v if v else None 
    
    