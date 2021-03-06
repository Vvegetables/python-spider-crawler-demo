#coding=utf-8

from collections import OrderedDict
import datetime
import os
import time
from urlparse import urljoin

from lxml import etree
from openpyxl.reader.excel import load_workbook
from openpyxl.workbook.workbook import Workbook
import requests
from sqlalchemy.engine import create_engine
from sqlalchemy.ext.declarative.api import declarative_base
from sqlalchemy.orm.session import sessionmaker
from sqlalchemy.sql.schema import Column
from sqlalchemy.sql.sqltypes import Integer, String, Text, DateTime

import MySQLdb as mysqldb
from self_mysql.mysql_connect import MysqlConnect


#'192.168.2.212','hanyj','Ihad#kd1234','CrawlerDB'
# LOGIN = dict(
#     username = 'hanyj',
#     password = 'Ihad#kd1234',
#     dbname = 'CrawlerDB',
#     host = '192.168.2.212',
#     port = '3306'
# )
LOGIN = dict(
    username = 'hanyj',
    password = 'CingHTa#1234',
    dbname = 'paTest',
    host = '118.31.168.208',
    port = '3306'
)

engine = create_engine('mysql://{username}:{password}@{host}:{port}/{dbname}?charset=utf8'.format(**LOGIN))

Base = declarative_base()

#college,major,direction,method,teacher,subject

class DoctorTable(Base): #User.__table__�鿴��Ϣ
    __tablename__ = 'professor_lists'
    id = Column('id',Integer,primary_key=True)
    entity = Column('entity',String(length=1000))
    college = Column('college',String(length=1000))
    major = Column('major',String(length=1000))
    direction = Column('direction',String(length=1000))
    method = Column('method',String(length=1000))
    teacher = Column('teacher',String(length=1000))
    subject = Column('subject',String(length=5000))
    
    
    def __repr__(self):
        return '{}'.format(self.id)

session = sessionmaker(bind=engine)
s = session()

wb = load_workbook("./folder/doctor_link.xlsx")

sheet = wb.get_sheet_by_name('Sheet1')

max_row = sheet.max_row
print max_row

main_urls = []
exit_flag = False

for i in range(1,max_row+1):
    if sheet.cell(row=i,column=1).value:
        main_urls.append((sheet.cell(row=i,column=1).value,sheet.cell(row=i,column=2).value))
        
headers = {
    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/49.0.2623.221 Safari/537.36 SE 2.X MetaSr 1.0',
    'Host': 'yz.chsi.com.cn',
    'Connection': 'keep-alive',
    'Cookie': 'JSESSIONID=5967E9A6C38A1E8CEF8D78AE341B815B; __utma=229973332.585229352.1531831545.1531836868.1531836868.1; __utmz=229973332.1531836868.1.1.utmcsr=(direct)|utmccn=(direct)|utmcmd=(none); aliyungf_tc=AQAAAJcZGSTJVwEAxjzOcwr0UzKAUqsL; acw_tc=AQAAAJUZ7EgrZwEAxjzOcwvgWsRDLh6z; _gat_gtag_UA_100524_7=1; _ga=GA1.3.585229352.1531831545; _gid=GA1.3.195840587.1531831545',
    'Refer' : 'http://yz.chsi.com.cn/bsmlcx/query.do?method=queryD',
    'Upgrade-Insecure-Requests':'1',
    'Cache-Control':'max-age=0',
}        

# proxies = {
#     'http' : '180.76.111.69:3128'
# }


ths_dict = {
    '院系所(招生人数)':'college',
    '专业(招生人数)':'major',
    '研究方向(招生人数)':'direction',
    '学习方式':'method',
    '指导教师(招生人数)':'teacher',
    '考试科目':'subject',
}

# s = requests.Session()

for _tuple in main_urls[250:]:
    if exit_flag:
        break
    
    retry = 2
    while retry:
        try:
            response = requests.get(url=_tuple[1],headers=headers,timeout=15)      #,proxies=proxies
            text = response.text
        except Exception,e:
            print str(e)
            retry -= 1
            time.sleep(3)
            continue
        break
    
    if not retry:
        with open('fail.log','a') as f:
            f.write('time:' + str(datetime.datetime.now()))
            f.write('fail:' + _tuple[1].encode('utf-8'))
            f.write('reason:timeout')
            f.write(os.linesep)
#             time.sleep(4)
        exit_flag = True
        continue
    
    html = etree.HTML(text)
    _table = html.xpath('//table[@id="result_bsml_table"]')#.extract_first()
    _trs = _table[0].xpath("//tr") if _table else []
    
    container = []
    
    if not _table:
        with open('fail.log','a') as f:
            f.write('time:' + str(datetime.datetime.now()))
            f.write('fail:' + _tuple[1].encode('utf-8'))
            f.write('reason:table is none')
            f.write(os.linesep)
            time.sleep(4)
        continue
    #题目
    ths = _table[0].xpath('//th')
    
    ths_list = []
    
    for i,th in enumerate(ths,0):
        ths_list.append(th.xpath('string(.)').strip().encode('utf-8'))
    
    td_nums = len(ths)
    _td_names = [0] * td_nums
    
    for _tr in _trs:
        tds = _tr.xpath('td')
        if not tds: #th
            continue
        temp = [0] * td_nums
#         e = 0
        ltd = len(tds)
        tds = ([0] * (td_nums-ltd)) + tds
        for i,td in enumerate(tds,0):
            if isinstance(td,int):  #补齐
                temp[i] = _td_names[i]
                continue
            try:    
                span = td.xpath('@rowspan')[0].strip()
            except IndexError:
                span = 1
                
            _title = td.xpath('string(.)').strip()
            
            if int(span) > 1:
                _td_names[i] = _title
                temp[i] = _title
            else:
                temp[i] = _title
                
        container.append({'entity':_tuple[0],ths_dict[ths_list[0]]:temp[0],ths_dict[ths_list[1]]:temp[1],ths_dict[ths_list[2]]:temp[2],ths_dict[ths_list[3]]:temp[3],ths_dict[ths_list[4]]:temp[4],ths_dict[ths_list[5]]:temp[5]})
        
#     _connect =MysqlConnect('192.168.2.212','hanyj','Ihad#kd1234','CrawlerDB') 
#     
#     _connect.batch_insert('educationnews', fields=('college','major','direction','method','teacher','subject'), values=container)
    
#     db = mysqldb.connect('192.168.2.212','hanyj','Ihad#kd1234','CrawlerDB',charset='utf8',port=3306)
#     cursor = db.cursor(cursorclass=mysqldb.cursors.DictCursor)
#     cursor.executemany("Insert into doctor_lib(college,major,direction,method,teacher,subject) values('%s','%s','%s','%s','%s','%s')",container)
        
    
    s.execute(DoctorTable.__table__.insert(),
              container)
    s.commit()
    
    time.sleep(6)
# session = sessionmaker(bind=engine)
# s = session()
# container = [{
#         'college':'c',
#         'major':'c',
#         'direction':'c',
#         'method':'c',
#         'teacher':'c',
#         'subject':'c',
#     }]        
# s.execute(DoctorTable.__table__.insert(),
#               container)        
# s.commit()        
#         
        