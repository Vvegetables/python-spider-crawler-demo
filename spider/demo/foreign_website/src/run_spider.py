#coding=utf-8
import logging
import time

from openpyxl.reader.excel import load_workbook
import redis

from spider.Schedule import Schedule
from spider.Task import Task
from utils.util import get_list_first_name
from utils.supervisor import Supervisor
from utils.url_redis import UrlRedis

SEQUENCE = 1

logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s %(filename)s[line:%(lineno)d] \
    %(levelname)s %(message)s',
    filename='Data/log.txt',
    )

def main():
#     print 1
#     client = redis.Redis()
#     list_first_name = get_list_first_name('Data/first_name.txt')
#     while True:
#         t = Task(client)
#         task = t.task
#         logging.info(task)
#         if task:
#             task.update(
#                 list_first_name=list_first_name,
#             )
#             process = Schedule(**task)
#             logging.info('{} spiding'.format(task['url']))
#             process.run()
#             t.save()
#         else:
#             time.sleep(50)
    
    wb = load_workbook('D:\\Coding\\python\\Console_project\\foreign_website\\src\\Data\\website.xlsx')
    sheet = wb.active
    maxrow = sheet.max_row
#     maxcolumn = sheet.max_column
#     maxrow = 10
    exit_param = {'exit':0}
    govern = Supervisor('/mnt/foreign_website/src/run_spider.py','/usr/bin/python /mnt/foreign_website/src/run_spider.py','/root/b.txt',exit_param)
    govern.start()
    for i in range(1,maxrow + 1):
        flag = UrlRedis.add_url(sheet.cell(row=i,column=3).value)
        if flag:
            continue
#         for j in range(1,maxcolumn + 1):
#             sheettitle = sheet.cell(row=1,column=j).value
#         sheetdata = sheet.cell(row=i,column=1).value #
        task = dict(
                url=sheet.cell(row=i,column=3).value,
                university=sheet.cell(row=i,column=2).value,
                region=sheet.cell(row=i,column=5).value,
                en_university=sheet.cell(row=i,column=4).value,
            ) 
#     task = dict(
#             url='http://www.ucl.ac.uk/news/all-news',
#             university='伦敦大学学院',
# #             department='城市与环境学院',
#             region = '英国',
#             en_university='University College London',
#         )
#         logging.info(task)
        if task:
    #         task.update(
    #             list_first_name=list_first_name,
    #         )
            process = Schedule(**task)
            logging.info('{} spiding'.format(task['url']))
            process.run()
    
    exit_param['exit'] = 1
    govern.join()
    

if __name__ == '__main__':
    main()
